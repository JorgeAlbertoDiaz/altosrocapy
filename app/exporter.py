"""Export the estados-socios grid to Excel (.xlsx) or PDF."""

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

PDF_ICON_REPLACEMENTS = {"\u2705": "Si"}


def _grid_data(tree):
    """(headers, rows) currently visible in the Treeview."""
    cols = [tree.heading(c)["text"] for c in tree["columns"]]
    rows = []
    for item in tree.get_children():
        rows.append([str(v) for v in tree.item(item, "values")])
    return cols, rows


def export_excel(tree, filepath) -> int:
    cols, rows = _grid_data(tree)
    wb = Workbook()
    ws = wb.active
    ws.title = "Estados Socios"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1A2430")
    ws.append(cols)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for r in rows:
        ws.append(r)
    # Reasonable column widths
    for idx, col in enumerate(cols, start=1):
        width = max([len(col)] + [len(r[idx - 1]) for r in rows]) + 4 if rows else len(col) + 4
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(width, 45)
    wb.save(filepath)
    return len(rows)


def export_pdf(tree, filepath) -> int:
    cols, rows = _grid_data(tree)
    # Portrait A4: usable width ~190 mm / 10 columns -> very small fonts.
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 8, "Estados de Socios - Altos Roca", align="C")
    pdf.ln(10)

    col_w = _pdf_col_widths(pdf, cols, rows)
    pdf.set_font("Helvetica", "B", 6)
    pdf.set_fill_color(26, 36, 48)
    pdf.set_text_color(255)
    for c, w in zip(cols, col_w):
        pdf.cell(w, 5, _pdf_safe(c), border=1, fill=True, align="C")
    pdf.ln()

    # Plain white rows: striping makes a table this dense unreadable.
    pdf.set_text_color(0)
    pdf.set_font("Helvetica", "", 5.5)
    for r in rows:
        for v, w in zip(r, col_w):
            pdf.cell(w, 4.5, _pdf_safe(v), border=1)
        pdf.ln()
    pdf.output(filepath)
    return len(rows)


def _pdf_col_widths(pdf, cols, rows):
    """Column widths proportional to expected content, fitting the page."""
    usable = pdf.w - pdf.l_margin - pdf.r_margin
    weights = [1.0, 2.8, 1.7, 0.8, 1.5, 1.5, 1.0, 0.9, 1.6, 0.8]
    total = sum(weights)
    return [round(usable * wgt / total, 1) for wgt in weights]


def _pdf_safe(text: str) -> str:
    """Map non-latin-1 glyphs (fpdf core fonts are latin-1 only)."""
    for k, v in PDF_ICON_REPLACEMENTS.items():
        text = text.replace(k, v)
    return text.encode("latin-1", "replace").decode("latin-1")
