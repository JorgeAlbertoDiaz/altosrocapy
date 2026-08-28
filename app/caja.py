"""Ingresos / Caja — libro diario de movimientos económicos del gimnasio.

Centraliza:
  - Cobros de cuotas (tbPagos)
  - Cobros de deudas (tb_RegistroDeudas canceladas)
  - Ingresos manuales (tbIngresosGenerales)
  - Gastos (tbGastosGenerales)
  - Balance acumulado = ingresos (DEBE) - gastos (HABER), calculado, nunca almacenado

Estética: WinForms / VB.NET clásica. Lista cronológica (más antiguo arriba).
"""

import datetime
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from calendar import month_name

try:
    from tkcalendar import DateEntry
except ImportError:
    DateEntry = None

try:
    from app import db
except ImportError:
    import db

# ── Constants ─────────────────────────────────────────────────────────────

W, H = 1024, 560
BG = "#F0F0F0"
FG = "#000000"
ENTRY_BG = "#FFFFFF"
BTN_BLUE = "#3B6FA0"
BTN_BLUE_ACTIVE = "#2D5A85"
SEL_BG = "#0078D7"
FN = ("Helvetica", 9)
FN_B = ("Helvetica", 9, "bold")

# Verbatim month list from prompt
MESES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
         "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

# Forma de pago: real values from tb_TipoPago (id -> descripcion)
FORMA_PAGO = [("Todos", None), ("Efectivo", "1"), ("Transferencia", "2"),
              ("Debito", "3"), ("Credito", "4"), ("Otro", "5")]


# ── Helpers / Data access ─────────────────────────────────────────────────

def _get_con():
    return db.get_connection()


def _tipopago_map():
    conn = _get_con()
    try:
        return {
            str(r[0]): r[1] for r in conn.execute(
                "SELECT Id, Descripcion FROM tb_TipoPago").fetchall()
        }
    finally:
        conn.close()


def _forma_pago(id_tipo):
    m = _tipopago_map()
    return m.get(str(id_tipo), "") if id_tipo else ""


def _socio_nombre_dni(id_socio):
    conn = _get_con()
    try:
        r = conn.execute(
            "SELECT Apellidos, Nombres, Documento FROM tbSocios WHERE idSocio = ?",
            (str(id_socio),)).fetchone()
        if not r:
            return "", ""
        nombre = f"{(r[0] or '').upper()}, {(r[1] or '').upper()}".strip(", ")
        return nombre, r[2] or ""
    finally:
        conn.close()


def _fmt_ts(raw):
    if not raw:
        return ""
    s = str(raw).strip()
    dt = None
    # Fecha con hora (sufijo .000 opcional) o solo fecha
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.datetime.strptime(s[:19], fmt)
            break
        except ValueError:
            continue
    if dt is None:
        return s
    if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
        return dt.strftime("%d/%m/%Y")
    return dt.strftime("%d/%m/%Y %H:%M")


def _dt_to_ts(d):
    return d.strftime("%Y-%m-%d %H:%M:%S.000")


def _collect_movements(fecha_d, fecha_h, usuario, id_tipo_pago,
                       incluir_ingresos, incluir_gastos):
    """Return a flat, unsorted list of movement dicts.

    Each dict: {'fecha_ts', 'id', 'detalle', 'debe', 'haber'}
      - debe  = money in (ingreso) -> positive in debe (suma)
      - haber = money out (gasto)  -> positive in haber
    """
    conn = _get_con()
    movs = []
    try:
        if incluir_ingresos:
            # 1) Cuotas cobradas (tbPagos) — debe
            rows = conn.execute(
                "SELECT p.idPago, p.idSocio, p.FechadePago, p.Importe, "
                "  p.UsuarioCobrador, p.idTipoPago, p.Observaciones "
                "FROM tbPagos p "
                "WHERE (p.Eliminado IS NULL OR p.Eliminado != '1') "
                "  AND substr(CAST(p.FechadePago AS TEXT), 1, 10) >= ? "
                "  AND substr(CAST(p.FechadePago AS TEXT), 1, 10) <= ?",
                (fecha_d[:10], fecha_h[:10])).fetchall()
            for r in rows:
                nombre, dni = _socio_nombre_dni(r[1])
                fp = _forma_pago(r[5])
                detalle = (f"Cuota Mensual - {nombre} - {dni}"
                           f"{' - ' + (r[6] or '') if r[6] else ''}"
                           f"{' - ' + (fp or '') if fp else ''}"
                           f" - Fec: {_fmt_ts(r[2])}")
                movs.append({
                    "fecha_ts": str(r[2]), "id": str(r[0]), "detalle": detalle,
                    "debe": float(r[3] or 0), "haber": 0.0,
                    "usuario": r[4] or "", "pago": str(r[5] or ""),
                })

            # 2) Deudas cobradas (canceladas) — debe
            rows = conn.execute(
                "SELECT d.idDeuda, d.idSocio, d.FechaCancelacion, "
                "  d.ImporteDeuda, d.Detalle, d.UsuarioCobrador "
                "FROM tb_RegistroDeudas d "
                "WHERE d.Cancelada = '1' "
                "  AND (d.Eliminado IS NULL OR d.Eliminado != '1') "
                "  AND substr(CAST(d.FechaCancelacion AS TEXT), 1, 10) >= ? "
                "  AND substr(CAST(d.FechaCancelacion AS TEXT), 1, 10) <= ?",
                (fecha_d[:10], fecha_h[:10])).fetchall()
            for r in rows:
                nombre, dni = _socio_nombre_dni(r[1])
                detalle = (f"Pago Deuda - {nombre} - {r[4] or ''}")
                movs.append({
                    "fecha_ts": str(r[2]), "id": str(r[0]), "detalle": detalle,
                    "debe": float(r[3] or 0), "haber": 0.0,
                    "usuario": r[5] or "", "pago": "0",
                })

            # 3) Ingresos generales — debe
            rows = conn.execute(
                "SELECT g.idIngreso, g.Detalle, g.Importe, g.Fecha, "
                "  g.UsuarioCobrador, g.idTipoPago "
                "FROM tbIngresosGenerales g "
                "WHERE (g.Eliminado IS NULL OR g.Eliminado != '1') "
                "  AND substr(CAST(g.Fecha AS TEXT), 1, 10) >= ? "
                "  AND substr(CAST(g.Fecha AS TEXT), 1, 10) <= ?",
                (fecha_d[:10], fecha_h[:10])).fetchall()
            for r in rows:
                detalle = f"Ingreso Extra - {r[1] or ''}"
                movs.append({
                    "fecha_ts": str(r[3]), "id": str(r[0]), "detalle": detalle,
                    "debe": float(r[2] or 0), "haber": 0.0,
                    "usuario": r[4] or "", "pago": str(r[5] or ""),
                })

        if incluir_gastos:
            # 4) Gastos — haber
            rows = conn.execute(
                "SELECT g.idGastos, g.Detalle, g.Importe, g.Fecha, "
                "  g.Usuario, g.idTipoPago "
                "FROM tbGastosGenerales g "
                "WHERE (g.Eliminado IS NULL OR g.Eliminado != '1') "
                "  AND substr(CAST(g.Fecha AS TEXT), 1, 10) >= ? "
                "  AND substr(CAST(g.Fecha AS TEXT), 1, 10) <= ?",
                (fecha_d[:10], fecha_h[:10])).fetchall()
            for r in rows:
                detalle = f"Gasto - {r[1] or ''}"
                movs.append({
                    "fecha_ts": str(r[3]), "id": str(r[0]), "detalle": detalle,
                    "debe": 0.0, "haber": float(r[2] or 0),
                    "usuario": r[4] or "", "pago": str(r[5] or ""),
                })
    finally:
        conn.close()

    # Apply user + forma de pago filters
    filtered = []
    for m in movs:
        if usuario and m["usuario"] != usuario:
            continue
        if id_tipo_pago and m["pago"] != id_tipo_pago:
            continue
        filtered.append(m)

    # Sort chronological, oldest first. Stable by fecha, then id.
    filtered.sort(key=lambda m: (m["fecha_ts"], m["id"]))
    return filtered


# ── Excel export ──────────────────────────────────────────────────────────

def _export_excel(fila, movs, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Caja"

    header_fill = PatternFill("solid", fgColor="3B6FA0")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["ID", "Detalle", "Debe", "Haber", "Saldo", "Fecha"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for m, saldo in zip(movs, fila):
        ws.append([m["id"], m["detalle"],
                   m["debe"] if m["debe"] else None,
                   m["haber"] if m["haber"] else None,
                   saldo, _fmt_ts(m["fecha_ts"])])

    widths = [10, 90, 12, 12, 14, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(path)


# ── Main Window ───────────────────────────────────────────────────────────

class CajaWindow(tk.Toplevel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.title("Ingresos / Caja")
        self.geometry(f"{W}x{H}")
        self.resizable(False, False)
        self.configure(bg=BG)
        self.bind("<Escape>", lambda _: self.destroy())

        self._movs = []
        self._saldo_acum = []
        self._build()
        # Initially set input states + load data (modo diaria por defecto)
        self._toggle_mode()

    def _build(self):
        # === FILTER BAR (no groupbox) ===
        frm = tk.Frame(self, bg=BG)
        frm.place(x=15, y=15, width=980, height=85)

        # --- Caja Diaria (radio, selected by default) ---
        self.diaria_var = tk.StringVar(value="diaria")
        tk.Radiobutton(frm, text="Caja Diaria", variable=self.diaria_var,
                       value="diaria", bg=BG, font=FN, fg=FG,
                       command=self._toggle_mode).place(x=0, y=0)
        self.fecha_var = tk.StringVar(value=datetime.date.today())
        if DateEntry is not None:
            self.dt_fecha = DateEntry(frm, width=22, background="#3B6FA0",
                                      foreground="white", borderwidth=1,
                                      date_pattern="dd/mm/yyyy", font=FN)
            self.dt_fecha.place(x=0, y=25, width=230, height=25)
        else:
            self.dt_fecha = None
            self.entry_fecha = tk.Entry(
                frm, textvariable=self.fecha_var, bg=ENTRY_BG, fg=FG,
                font=FN, relief="solid", bd=1)
            self.entry_fecha.place(x=0, y=25, width=230, height=25)

        # --- Filtrar por Período Mensual (radio) ---
        tk.Radiobutton(frm, text="Filtrar por Período Mensual",
                       variable=self.diaria_var, value="mensual",
                       bg=BG, font=FN, fg=FG,
                       command=self._toggle_mode).place(x=250, y=0)

        self.anio_var = tk.StringVar(value=str(datetime.date.today().year))
        self.spin_anio = tk.Spinbox(
            frm, from_=2000, to=2100, textvariable=self.anio_var,
            width=5, font=FN, command=self._on_change, increment=1)
        self.spin_anio.place(x=250, y=25, width=70, height=25)
        self.anio_var.trace_add("write", lambda *_: self._on_change())

        self.mes_var = tk.StringVar(value=MESES[datetime.date.today().month - 1])
        self.cb_mes = ttk.Combobox(frm, values=MESES, textvariable=self.mes_var,
                                   state="readonly", font=FN, width=12)
        self.cb_mes.place(x=330, y=25, width=110, height=25)
        self.cb_mes.bind("<<ComboboxSelected>>", lambda _: self._on_change())

        # --- Filtro Usuario ---
        tk.Label(frm, text="Usuario", bg=BG, fg=FG, font=FN).place(x=460, y=0)
        self.usuario_var = tk.StringVar(value="Todos")
        usuarios = ["Todos"] + self._lista_usuarios()
        cb_usu = ttk.Combobox(frm, values=usuarios, textvariable=self.usuario_var,
                              state="readonly", font=FN, width=12)
        cb_usu.place(x=460, y=25, width=100, height=25)
        cb_usu.bind("<<ComboboxSelected>>", lambda _: self._on_change())

        # --- Forma de Pago del Plan ---
        tk.Label(frm, text="Forma de Pago del Plan", bg=BG, fg=FG,
                 font=FN).place(x=585, y=0)
        self.pago_var = tk.StringVar(value="Todos")
        cb_pago = ttk.Combobox(frm, values=[p[0] for p in FORMA_PAGO],
                               textvariable=self.pago_var,
                               state="readonly", font=FN, width=14)
        cb_pago.place(x=585, y=25, width=120, height=25)
        cb_pago.bind("<<ComboboxSelected>>", lambda _: self._on_change())

        # --- Checkboxes (DERECHA) ---
        self.incluir_ingr_var = tk.BooleanVar(value=True)
        tk.Checkbutton(frm, text="INCLUIR INGRESOS", variable=self.incluir_ingr_var,
                       bg=BG, font=FN, command=self._on_change).place(x=745, y=8)

        self.incluir_gast_var = tk.BooleanVar(value=True)
        tk.Checkbutton(frm, text="INCLUIR GASTOS", variable=self.incluir_gast_var,
                       bg=BG, font=FN, command=self._on_change).place(x=745, y=38)

        # --- Actualizar datos (esquina superior derecha) ---
        tk.Button(frm, text="Actualizar datos", bg=BTN_BLUE, fg="#FFF",
                  font=FN_B, relief="flat", activebackground=BTN_BLUE_ACTIVE,
                  cursor="hand2", command=self._refresh).place(x=870, y=32,
                                                               width=110,
                                                               height=25)

        # === MAIN GRID (990x305) ===
        frm_grid = tk.Frame(self, bg="#E8E8E8")
        frm_grid.place(x=15, y=110, width=990, height=305)

        cols = ("id", "detalle", "debe", "haber", "saldo")
        self.tree = ttk.Treeview(frm_grid, columns=cols, show="headings",
                                 selectmode="browse")
        self.tree.heading("id", text="ID")
        self.tree.heading("detalle", text="Detalle")
        self.tree.heading("debe", text="Debe")
        self.tree.heading("haber", text="Haber")
        self.tree.heading("saldo", text="Saldo")
        self.tree.column("id", width=50, anchor="center", stretch=False)
        self.tree.column("detalle", width=460, stretch=True)
        self.tree.column("debe", width=120, anchor="e", stretch=False)
        self.tree.column("haber", width=120, anchor="e", stretch=False)
        self.tree.column("saldo", width=140, anchor="e", stretch=False)

        style = ttk.Style(self)
        style.configure("CJ.Treeview", rowheight=22, font=FN,
                        background="#FFF", fieldbackground="#FFF")
        style.configure("CJ.Treeview.Heading", font=FN_B)
        style.map("CJ.Treeview",
                  background=[("selected", SEL_BG)],
                  foreground=[("selected", "#FFF")])
        self.tree.configure(style="CJ.Treeview")

        vsb = ttk.Scrollbar(frm_grid, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frm_grid, orient="horizontal",
                            command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.place(x=0, y=0, relwidth=0.985, relheight=0.96)
        vsb.place(relx=0.985, y=0, relheight=0.96)
        hsb.place(x=0, rely=0.96, relwidth=0.985)

        # Procesando overlay (shown while filtering/loading)
        self.lbl_proc = tk.Label(
            frm_grid, text="⏳ Procesando reporte...", bg="#FFFBCC",
            fg="#7A5B00", font=("Helvetica", 14, "bold"), relief="solid", bd=1,
        )
        self.lbl_proc.place(relx=0.5, rely=0.5, anchor="center")
        self.lbl_proc.place_forget()

        # === BOTTOM PANEL (990x60) ===
        frm_bot = tk.Frame(self, bg=BG)
        frm_bot.place(x=15, y=430, width=990, height=60)

        tk.Label(frm_bot, text="CAJA:", bg=BG, fg=FG, font=FN_B).place(x=0, y=15)

        tk.Button(frm_bot, text="Ver Importe", bg=BTN_BLUE, fg="#FFF",
                  font=FN_B, relief="flat", activebackground=BTN_BLUE_ACTIVE,
                  cursor="hand2",
                  command=self._ver_importe).place(x=60, y=8, width=95,
                                                   height=25)

        tk.Button(frm_bot, text="Exportar a Excel", bg=BTN_BLUE, fg="#FFF",
                  font=FN_B, relief="flat", activebackground=BTN_BLUE_ACTIVE,
                  cursor="hand2",
                  command=self._exportar_excel).place(x=745, y=8, width=125,
                                                      height=25)

        tk.Button(frm_bot, text="Salir", bg="#D9D9D9", fg="#333333", font=FN_B,
                  relief="flat", activebackground="#BFBFBF", cursor="hand2",
                  command=self.destroy).place(x=915, y=8, width=75, height=25)

    # ── Filter helpers ────────────────────────────────────────────────────

    def _lista_usuarios(self):
        conn = _get_con()
        try:
            s = set()
            for r in conn.execute(
                "SELECT UsuarioCobrador FROM tbPagos WHERE UsuarioCobrador IS NOT NULL"):
                if str(r[0]).strip() != "-------":
                    s.add(str(r[0]).strip())
            for r in conn.execute(
                "SELECT Usuario FROM tbGastosGenerales WHERE Usuario IS NOT NULL"):
                if str(r[0]).strip() != "-------":
                    s.add(str(r[0]).strip())
            for r in conn.execute(
                "SELECT UsuarioCobrador FROM tbIngresosGenerales WHERE UsuarioCobrador IS NOT NULL"):
                if str(r[0]).strip() != "-------":
                    s.add(str(r[0]).strip())
            return sorted(s)
        finally:
            conn.close()

    def _toggle_mode(self):
        self._on_change()

    def _on_change(self):
        self._refresh()

    def _rango(self):
        """Return (fecha_d, fecha_h) TS strings based on selected mode."""
        if self.diaria_var.get() == "diaria":
            if self.dt_fecha is not None:
                d = self.dt_fecha.get_date()
            else:
                parts = self.fecha_var.get().split("/")
                d = datetime.date(int(parts[2]), int(parts[1]), int(parts[0]))
            return (f"{d:%Y-%m-%d} 00:00:00", f"{d:%Y-%m-%d} 23:59:59")
        else:
            anio = int(self.anio_var.get())
            mes = MESES.index(self.mes_var.get()) + 1
            import calendar
            last = calendar.monthrange(anio, mes)[1]
            return (f"{anio:04d}-{mes:02d}-01 00:00:00",
                    f"{anio:04d}-{mes:02d}-{last:02d} 23:59:59")

    def _filtro_pago(self):
        pago = self.pago_var.get()
        for label, pid in FORMA_PAGO:
            if label == pago:
                return pid
        return None

    # ── Refresh ───────────────────────────────────────────────────────────

    def _toggle_mode(self):
        # Disable the inputs that don't belong to the active mode.
        if self.diaria_var.get() == "diaria":
            # Diaria: datepicker enabled, period (año/mes) disabled
            if self.dt_fecha is not None:
                self.dt_fecha.configure(state="normal")
            else:
                self.entry_fecha.configure(state="normal")
            self.spin_anio.configure(state="disabled")
            self.cb_mes.configure(state="disabled")
        else:
            # Mensual: period enabled, datepicker disabled
            if self.dt_fecha is not None:
                self.dt_fecha.configure(state="disabled")
            else:
                self.entry_fecha.configure(state="disabled")
            self.spin_anio.configure(state="normal")
            self.cb_mes.configure(state="readonly")
        self._refresh()

    def _refresh(self):
        fecha_d, fecha_h = self._rango()
        usuario = self.usuario_var.get()
        usuario = None if usuario == "Todos" else usuario
        id_tipo_pago = self._filtro_pago()
        incl_ingr = self.incluir_ingr_var.get()
        incl_gast = self.incluir_gast_var.get()

        # Show loading overlay + disable re-entry
        self.lbl_proc.place(relx=0.5, rely=0.5, anchor="center")
        self.tree.configure(cursor="watch")
        self.update_idletasks()

        def _worker():
            try:
                movs = _collect_movements(
                    fecha_d, fecha_h, usuario, id_tipo_pago,
                    incl_ingr, incl_gast)
            except Exception as e:
                self.after(0, lambda: self._apply_worker_error(e))
                return
            self.after(0, lambda: self._apply_movements(movs))

        threading.Thread(target=_worker, daemon=True).start()

    def _apply_movements(self, movs):
        self._movs = movs
        self.tree.delete(*self.tree.get_children())

        saldo = 0.0
        self._saldo_acum = []
        for m in self._movs:
            saldo += m["debe"] - m["haber"]
            self._saldo_acum.append(saldo)
            self.tree.insert("", "end", values=(
                m["id"], m["detalle"],
                f"{m['debe']:.0f}" if m["debe"] else "",
                f"{m['haber']:.0f}" if m["haber"] else "",
                f"{saldo:.0f}",
            ))
        self._total_actual = saldo

        self.lbl_proc.place_forget()
        self.tree.configure(cursor="")

    def _apply_worker_error(self, e):
        self.lbl_proc.place_forget()
        self.tree.configure(cursor="")
        messagebox.showerror("Error", f"Error al procesar el reporte: {e}",
                             parent=self)

    def _ver_importe(self):
        total = self._saldo_acum[-1] if self._saldo_acum else 0.0
        top = tk.Toplevel(self)
        top.title("CAJA ACTUAL")
        top.geometry("280x90")
        top.resizable(False, False)
        top.configure(bg=BG)
        tk.Label(top, text="CAJA ACTUAL", bg=BG, fg=FG, font=FN_B).pack(pady=8)
        tk.Label(top, text=f"${total:.0f}", bg=BG, fg="#003399",
                 font=("Helvetica", 18, "bold")).pack()
        top.transient(self)
        top.grab_set()

    def _exportar_excel(self):
        fecha_d, fecha_h = self._rango()
        if not self._movs:
            messagebox.showinfo("Exportar", "No hay datos para exportar.", parent=self)
            return
        path = filedialog.asksaveasfilename(
            parent=self, defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"caja_{fecha_d[:10]}_a_{fecha_h[:10]}.xlsx",
            title="Exportar a Excel")
        if not path:
            return
        try:
            _export_excel(self._saldo_acum, self._movs, path)
            messagebox.showinfo("Exportar", f"Exportado a:\n{path}", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}", parent=self)


def open_window(parent=None):
    return CajaWindow(parent)


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    open_window(root)
    root.mainloop()
